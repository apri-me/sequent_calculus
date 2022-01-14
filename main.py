from textwrap import fill
from sequent import Sequent
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill

class SequentTreePrinter:
    sequent_root: Sequent
    color: str

    def __init__(self, sequent_root: Sequent, is_proved) -> None:
        if not sequent_root.is_calculated == True:
            raise SequentNotCalculatedException()
        self.sequent_root = sequent_root
        self.color = "0042f569" if is_proved else "00f55142"

    def print_in_workbook(self):
        wb = Workbook()
        sheet = wb.active
        self.generate_sheet(sheet)
        wb.close()
        wb.save("SequentTree.xlsx")
    
    def generate_sheet(self, sheet):
        self.sequent_root.cal_size()
        self.sequent_root.cal_children_space_before()
        max_length = self.sequent_root.cal_max_length()
        self.put_sequent_in_sheet(self.sequent_root, max_length, sheet, self.color)
    
    @classmethod
    def put_sequent_in_sheet(cls, sequent: Sequent, max_length, sheet, color):
        if sequent.size > 1:
            sheet.merge_cells(
                                start_row=sequent.level, end_row=sequent.level, 
                                start_column=sequent.space_before+1, 
                                end_column=sequent.space_before+sequent.size
                            )
        cell = sheet.cell(sequent.level, sequent.space_before+1)
        cell.alignment = Alignment(horizontal='center', vertical="center")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.value = str(sequent)
        if not sequent.children:
            sheet.column_dimensions[cell.column_letter].width = int(3 * len(str(sequent)))
        for ch in sequent.children:
            cls.put_sequent_in_sheet(ch, max_length, sheet, color)


class SequentNotCalculatedException(Exception):
    pass

print("""***Welcome to Sequent Calculus Program.***

Try to write your formula below. 
Note: You can use connectives like 'and' by '/\\', 'or' by '\\/', 'negation' by '~' and 'imply' by '->'.
Note: Surround all of your formula with parens to make it well-formed!
Note: The sequent tree will be outputed in file 'SequentTree.xlsx'!
Note: If the formula has been proved, xl file will be shown by green background and otherwise it will be red.
Note: Use spaces between connectives and atomics. It will output wrong if you don't do this.
""")
root = Sequent.init_by_formula(input("Type your formula: "))
is_proved = root.calculate()
printer = SequentTreePrinter(root, is_proved)
printer.print_in_workbook()